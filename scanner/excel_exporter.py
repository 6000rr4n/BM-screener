"""Excel exporter: write scan results to XLSX using openpyxl."""

import logging
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def export_to_excel(signals: list[dict[str, Any]], filepath: str) -> str:
    """Export signal dicts to a formatted XLSX file.

    Returns the filepath written.
    """
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Scanner Results"

    # Headers
    headers = [
        "Ticker", "Price", "Signal Type", "Direction", "Stock Strategy",
        "Option Strategy", "DTE", "Regime", "Score", "RSI",
        "BBW%", "ATR%", "IV/HV%", "IV Rank", "Weekly Trend",
        "Weekly RSI", "TTM Squeeze", "TTM Direction", "TTM Color",
        "Confirmed", "Patterns", "Divergence",
    ]

    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(bold=True, color="f1f5f9", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    green_font = Font(color="22c55e")
    red_font = Font(color="ef4444")

    for row_idx, sig in enumerate(signals, 2):
        row_data = [
            sig.get("ticker", ""),
            sig.get("last_price", 0),
            sig.get("signal_type", ""),
            sig.get("signal_direction", ""),
            "BUY" if sig.get("signal_direction") == "BULLISH" else "SELL",
            sig.get("option_strategy", ""),
            sig.get("dte_range", ""),
            sig.get("regime", ""),
            sig.get("score", 0),
            sig.get("rsi"),
            sig.get("bbw_pct"),
            sig.get("atr_pct"),
            sig.get("iv"),
            sig.get("iv_rank"),
            sig.get("weekly_trend", "N/A"),
            sig.get("weekly_rsi"),
            "ON" if sig.get("ttm_squeeze_on") else ("FIRED" if sig.get("ttm_squeeze_off") else "OFF"),
            sig.get("ttm_momentum_dir", "FLAT"),
            sig.get("ttm_hist_color", ""),
            "Yes" if sig.get("confirmed") else "No",
            sig.get("patterns", ""),
            "Yes" if sig.get("divergence") else "No",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

        # Color the direction cell
        direction_cell = ws.cell(row=row_idx, column=4)
        if sig.get("signal_direction") == "BULLISH":
            direction_cell.font = green_font
        else:
            direction_cell.font = red_font

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(signals) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_width = max(max_width, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_width + 2, 25)

    wb.save(filepath)
    logger.info(f"Excel exported: {filepath} ({len(signals)} signals)")
    return filepath
